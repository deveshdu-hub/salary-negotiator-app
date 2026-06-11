import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Initialize workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "North Div Project Tracker"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Define premium color palette (Corporate Navy Theme)
COLOR_PRIMARY = "1A365D"    # Deep Navy for main title
COLOR_HEADER = "2A4365"     # Slate Blue for table headers
COLOR_TEXT_LIGHT = "FFFFFF" # White text
COLOR_ZEBRA = "F7FAFC"      # Very light grey for alternate rows
COLOR_BORDER = "E2E8F0"     # Soft grey for boundaries

# Status tracking conditional colors
FILL_UPCOMING = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid") # Soft Blue
FILL_ONGOING = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")  # Soft Yellow
FILL_COMPLETION = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Soft Teal

# Styles
font_title = Font(name="Segoe UI", size=16, bold=True, color=COLOR_TEXT_LIGHT)
font_header = Font(name="Segoe UI", size=10, bold=True, color=COLOR_TEXT_LIGHT)
font_data = Font(name="Segoe UI", size=10, bold=False, color="2D3748")
font_data_bold = Font(name="Segoe UI", size=10, bold=True, color="1A365D")

fill_title = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
fill_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

# 1. Create Main Title Banner
ws.merge_cells("A1:P1")
title_cell = ws["A1"]
title_cell.value = "BERGER PROTECTON (ADMIXTURE) - NORTH DIVISION MASTER TRACKER"
title_cell.font = font_title
title_cell.fill = fill_title
title_cell.alignment = align_center
ws.row_dimensions[1].height = 40

# 2. Define Headers
headers = [
    "Project ID", "State / Hub", "Client Category", "Project Name", "Lifecycle Stage",
    "Primary EPC Contractor", "Key Decision Maker (Name/Role/Contact)", "Structural Consultant (Spec Authority)",
    "QC / Plant Lead (Trial Authority)", "Incumbent Competitor Brand", "Target Application Zone",
    "Berger Counterweapon Product", "Current Action Status / Latest Update", "Next Concrete Action Required",
    "Target Date", "Win Probability (%)"
]

# Write Headers
ws.row_dimensions[2].height = 28
for col_num, header_text in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = header_text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# 3. Insert Baseline Master Data
mock_data = [
    ["PRJ-01", "Delhi-NCR", "Metro Rail", "Delhi Metro Phase IV (Golden Line)", "Upcoming", "L&T Construction", "Arjun Mehta (Procurement Head)", "DMRC Design Board", "S. Sharma (QC Manager)", "Sika India", "Underground Cut-&-Cover Tunnels", "ProHyperplast SP & HS ProCrystal 100", "Bidding stage active. Structural drawing pulled.", "Schedule technical meeting with DMRC Consultant for spec-in", "2026-07-15", "65%"],
    ["PRJ-02", "Uttar Pradesh", "NHAI / Expressways", "Ganga Expressway (Phase 2 Pours)", "Ongoing", "PNC Infratech", "V. K. Singh (Project Director)", "L N Malviya Infra", "R. Chaudhary (Plant QC)", "Fosroc India", "Mass road beds & bridge decks", "ProSuperplast RT", "Trial mix requested due to slump loss complaints in summer heat.", "Deliver product samples to site batching yard for initial trial mix", "2026-06-25", "80%"],
    ["PRJ-03", "Delhi-NCR", "Mega Private / Commercial", "DLF Cybercity Phase 2 Expansion", "Upcoming", "Tata Projects", "Rajesh Kapoor (VP Infrastructure)", "Mantec Consultants", "Amit Pal (Site In-charge)", "MC-Bauchemie", "Deep basement rafts & structural foundation piles", "Hs ProCrystal 100 & HS ProCem CI", "Architectural blueprint finalized. Sub-surface parameters mapped.", "Pitch crystalline integration benefits directly to Mantec Lead", "2026-07-20", "50%"],
    ["PRJ-04", "Jammu & Kashmir", "Metro Rail", "Jammu Metro Neo Elevated viaducts", "Upcoming", "Designated Concessionaire", "TBD (Tender Stage)", "RITES Limited", "TBD", "CAC (Additives)", "Elevated Viaduct Precast Segments", "HS ProCem 9000 AB/PC", "Casting yard layout approvals active.", "Submit technical datasheet of HS ProCem 9000 to RITES approval board", "2026-08-10", "70%"],
    ["PRJ-05", "Uttar Pradesh", "Metro Rail", "Agra Metro Underground Corridors", "Ongoing", "Afcons Infrastructure", "S. Mukherjee (Project Head)", "UPMRC Technical Cell", "K. Dwivedi (QC Lead)", "Sika India", "Diaphragm walls & tunnel segment casting", "ProHyperplast series", "Continuous pours active. Incumbent experiencing mild bleeding issues.", "Execute on-site comparison mix demonstrating superior cohesion", "2026-06-30", "75%"],
    ["PRJ-06", "Punjab", "PWD / Civic Utilities", "Amritsar Smart Parking Complex", "Completion Stage", "Local Grade-A Contractor", "G. S. Dhillon (Executive Engineer)", "Punjab PWD Design Cell", "Harpreet Singh (Site Lead)", "Dr. Fixit (Pidilite)", "Finished structural piers & multi-level basement decks", "Protecton Coatings & Expansion Joint Sealants", "Core structural framework handovers complete. Moving to finishing.", "Present anti-carbonation coatings and PU flooring specifications", "2026-07-05", "85%"]
]

# Write Data and Apply Formatting
for row_idx, row_data in enumerate(mock_data, 3):
    ws.row_dimensions[row_idx].height = 24
    is_even = (row_idx % 2 == 0)
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = font_data
        cell.border = thin_border
        
        # Zebra striping base background
        if is_even:
            cell.fill = fill_zebra
            
        # Left align text descriptions, center align metrics/IDs
        if col_idx in [1, 2, 3, 5, 15, 16]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
            
        # Highlight Project IDs
        if col_idx == 1:
            cell.font = font_data_bold
            
        # Highlight Lifecycle Stage contextually
        if col_idx == 5:
            cell.font = font_data_bold
            if value == "Upcoming":
                cell.fill = FILL_UPCOMING
            elif value == "Ongoing":
                cell.fill = FILL_ONGOING
            elif value == "Completion Stage":
                cell.fill = FILL_COMPLETION

# 4. Auto-fit column widths elegantly with an extra padding cushion
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    # Don't size based on row 1 (merged banner)
    for cell in col:
        if cell.row > 1 and cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save Workbook
output_filename = "Berger_North_Division_Daily_Tracker.xlsx"
wb.save(output_filename)
print(f"Master Sheet generated successfully: '{output_filename}'")
